#!/usr/bin/env python3
# MIT License
#
# generate-documents.py
#
# Copyright (c) 2011-2026 Henning Rauch
#
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:
#
# The above copyright notice and this permission notice shall be included in all
# copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
#
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
# SOFTWARE.

"""Authoring tool for the knowledge-demo documents (feature knowledge-demo).

Run BY HAND when the content changes, not as part of any build:

    pip install reportlab openpyxl matplotlib
    python samples/documents/generate-documents.py

It writes the three documents the Samples gallery ingests live:

    nw-rca-wtg-a17.pdf    a root-cause analysis carrying a figure   (docling PDF path)
    nw-fleet-register.xlsx  the maintenance register table          (docling XLSX path)
    nw-std-0417.md        an engineering standard, text only        (no sidecar at all)

The outputs are COMMITTED. The sample build is vite-node TypeScript, and adding a PDF and
spreadsheet toolchain to it just to re-derive three static files on every run is bloat; the
repo already pins other sample inputs (the stored SBOM, the curated movie list) instead of
refetching them.

Three constraints the content must respect, each verified against a live instance and
recorded in features/*/knowledge-demo/spec.md:

1. Chunking merges sections below 800 characters, so every section here is deliberately
   longer than that. Short sections collapse the whole document into one chunk and the
   chunk chain disappears.
2. Asset tags come from windFarmFleet.json and nowhere else. Structural linking is
   ordinal-exact, so a single character of drift links nothing at all.
3. A figure CAPTION does not survive conversion (docling attaches it to the picture, which
   the chunker does not model). Anything the figure means is therefore stated in body prose
   as well. That is also just good technical writing.
"""

from __future__ import annotations

import datetime as dt
import json
import os
import re
import tempfile
import zipfile
from pathlib import Path

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.enums import TA_JUSTIFY
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer

HERE = Path(__file__).resolve().parent
FLEET_PATH = HERE.parent.parent / "fallen-8-web-ui" / "scripts" / "samples" / "data" / "windFarmFleet.json"

# Figure ink: near-black for the DATA, brand red reserved for the alarm LIMIT (a status
# meaning, carried with a dash pattern and a text label so it never relies on colour alone).
INK = "#1F2933"
LIMIT = "#E2001A"
GRID = "#D6D8DA"

# Fixed timestamp so regenerating does not churn the committed files. Setting the OOXML core
# properties is NOT sufficient for that: openpyxl stamps every zip member with time.localtime()
# at save, so the archive churns even when the content is identical. normalize_xlsx() below
# rewrites them, which is what actually makes the .xlsx reproducible.
FIXED_TIME = dt.datetime(2026, 3, 20, 9, 0, 0)
ZIP_TIME = (2026, 3, 20, 9, 0, 0)


def load_fleet() -> dict:
    with open(FLEET_PATH, encoding="utf-8") as handle:
        return json.load(handle)


def normalize_xlsx(path: Path) -> None:
    """Rewrites an OOXML archive so its bytes are reproducible.

    Two independent sources of churn, both verified by regenerating twice and diffing:
    every zip entry carries the wall-clock time it was written, AND openpyxl stamps
    `dcterms:modified` in docProps/core.xml with the current time at save, overriding the
    `wb.properties.modified` we set. Without both fixes the committed .xlsx shows up dirty in
    `git status` after a no-op regeneration.
    """
    stamp = FIXED_TIME.strftime("%Y-%m-%dT%H:%M:%SZ")
    with zipfile.ZipFile(path) as source:
        members = [(item, source.read(item.filename)) for item in source.infolist()]
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as target:
        for item, payload in members:
            if item.filename == "docProps/core.xml":
                text = payload.decode("utf-8")
                for field in ("created", "modified"):
                    text = re.sub(
                        rf"(<dcterms:{field}[^>]*>)[^<]*(</dcterms:{field}>)",
                        rf"\g<1>{stamp}\g<2>",
                        text,
                    )
                payload = text.encode("utf-8")
            fixed = zipfile.ZipInfo(item.filename, date_time=ZIP_TIME)
            fixed.compress_type = item.compress_type
            fixed.external_attr = item.external_attr
            target.writestr(fixed, payload)


# --------------------------------------------------------------------------------------
# The figure: a vibration spectrum showing the diagnostic sideband pattern.
# One series, so no legend (the title names it); recessive axes; selective direct labels.
# --------------------------------------------------------------------------------------
def build_figure(path: str, alarm: float) -> None:
    """The 12 March spectrum, which must SUPPORT the argument the prose makes.

    The narrative is that broadband energy was still modest while the sidebands were already
    unambiguous, so the mesh peak sits BELOW the broadband alarm line, not above it. Two things
    follow from taking that seriously: the alarm line is a broadband quantity and this is a
    narrowband spectrum, so it is labelled "for reference" rather than implying a single line can
    breach it; and no spectral line here may equal the 5.9 mm/s broadband reading, because
    broadband is the quadrature sum of the whole band and is necessarily larger than any one line.
    """
    mesh_hz = 312.0
    pass_hz = 14.2
    mesh_peak = 3.0
    freq = np.linspace(250.0, 375.0, 2400)

    def peak(centre: float, height: float, width: float) -> np.ndarray:
        return height * np.exp(-0.5 * ((freq - centre) / width) ** 2)

    # Broadband floor plus the mesh frequency plus symmetric sidebands at the bearing pass
    # frequency. Deterministic by construction: no PRNG anywhere.
    amplitude = 0.28 + 0.10 * np.sin(freq / 7.0) ** 2
    amplitude += peak(mesh_hz, mesh_peak, 1.4)
    for order in (1, 2, 3):
        height = 1.9 / order
        amplitude += peak(mesh_hz - order * pass_hz, height, 1.2)
        amplitude += peak(mesh_hz + order * pass_hz, height, 1.2)

    fig, ax = plt.subplots(figsize=(6.6, 2.9), dpi=200)
    ax.plot(freq, amplitude, color=INK, linewidth=2.0, solid_capstyle="round")
    ax.axhline(alarm, color=LIMIT, linewidth=1.6, linestyle=(0, (5, 3)))

    ax.annotate(
        f"{alarm} mm/s broadband alarm limit (for reference)",
        xy=(freq[0] + 3, alarm + 0.16),
        color=LIMIT,
        fontsize=7.5,
        fontweight="bold",
    )
    ax.annotate(
        f"gear mesh {mesh_hz:.0f} Hz",
        xy=(mesh_hz, mesh_peak),
        xytext=(mesh_hz + 18, mesh_peak + 0.55),
        color=INK,
        fontsize=7.5,
        arrowprops={"arrowstyle": "-", "color": INK, "linewidth": 0.8},
    )
    ax.annotate(
        f"sidebands spaced {pass_hz} Hz",
        xy=(mesh_hz - pass_hz, 1.9),
        xytext=(mesh_hz - 60, 2.7),
        color=INK,
        fontsize=7.5,
        arrowprops={"arrowstyle": "-", "color": INK, "linewidth": 0.8},
    )

    ax.set_xlabel("Frequency (Hz)", fontsize=8, color=INK)
    ax.set_ylabel("Velocity (mm/s RMS)", fontsize=8, color=INK)
    ax.set_ylim(0, 5.6)
    ax.set_xlim(freq[0], freq[-1])
    ax.grid(True, axis="y", color=GRID, linewidth=0.6)
    ax.set_axisbelow(True)
    for side in ("top", "right"):
        ax.spines[side].set_visible(False)
    for side in ("left", "bottom"):
        ax.spines[side].set_color(GRID)
    ax.tick_params(labelsize=7.5, colors=INK, length=3)

    fig.tight_layout()
    fig.savefig(path, metadata={"Software": None, "Creation Time": None})
    plt.close(fig)


# --------------------------------------------------------------------------------------
# Document 1: the root-cause analysis (PDF, carries the figure)
# --------------------------------------------------------------------------------------
def rca_sections(fleet: dict) -> list[tuple[str, str]]:
    std = fleet["standard"]
    suspect = next(b for b in fleet["batches"] if b["suspect"])
    subject = next(t for t in fleet["turbines"] if t["tag"] == "WTG_A17")
    gbx = subject["gearbox"]
    signer = next(t for t in fleet["technicians"] if t["signsDocuments"])
    site = next(s for s in fleet["sites"] if s["tag"] == subject["site"])
    alarm = std["alarmMmPerS"]

    return [
        (
            "Scope and summary",
            f"This root cause analysis covers the high speed stage failure of gearbox "
            f"{gbx['tag']}, installed in turbine {subject['tag']} at the {site['name']} site "
            f"operated by {fleet['operator']} in {fleet['region']}. The turbine sits on the "
            f"{subject['substation']} collector circuit. On 14 March 2026 the condition "
            f"monitoring system recorded a broadband velocity of {gbx['vibration']} mm/s RMS "
            f"at the high speed stage, above the {alarm} mm/s alarm limit that standard "
            f"{std['tag']} defines, and the unit was removed from service the same day. "
            f"Borescope inspection and subsequent strip down found advanced spalling on the "
            f"upwind planet bearing outer raceway. The gearbox was supplied by "
            f"{fleet['vendors'][0]} and the strip down was witnessed by their field engineers at "
            f"Esbjerg, with hardness coupons tested by an independent laboratory. "
            f"This report establishes the failure mechanism, explains why the alarm fired when "
            f"it did rather than earlier, and identifies the contributing factor that brought "
            f"the failure forward of design life. It was prepared and signed by "
            f"{signer['name']}, {signer['role'].lower()}, {fleet['operator']}. The findings "
            f"apply to one unit, but the contributing factor does not, which is the reason "
            f"this report ends with a fleet action rather than a single repair order.",
        ),
        (
            "Sequence of events",
            f"The unit ran without exception through the 2025 autumn campaign. Work order "
            f"{gbx['workOrder']} recorded a routine oil change and filter inspection on "
            f"{gbx['lastService']}, carried out by {gbx['technician']}; the filter debris was "
            f"noted as light and within limits at that time, which in hindsight was the first "
            f"observable symptom rather than a clean result. Through January 2026 the trended "
            f"velocity at the high speed stage rose slowly from 2.1 mm/s to 2.8 mm/s. That "
            f"rise stayed below the {std['warningMmPerS']} mm/s warning level, so it raised no "
            f"automatic flag, and because the increase was gradual it did not trigger a rate "
            f"of change notification either. In the last week of February the spectrum changed "
            f"character: the overall level barely moved, but discrete sidebands appeared around "
            f"the gear mesh frequency. On 11 March the overall level began to climb steeply, "
            f"reaching {gbx['vibration']} mm/s on 14 March, at which point the alarm fired and "
            f"the turbine was stopped. The elapsed time from the first clear spectral signature "
            f"to the alarm was about two weeks, and the elapsed time from the alarm to removal "
            f"was under twelve hours.",
        ),
        (
            "Why the bearing failed",
            # This is the section a reader lands on when they ask "why did the bearing fail",
            # so it deliberately names the unit as well as explaining the mechanism: the
            # identifiers are what link the explanation to the real asset in the graph.
            f"The mechanism established for {gbx['tag']} on strip down is subsurface initiated "
            f"rolling contact fatigue, progressing to macroscopic spalling. It runs in a "
            f"sequence, and each step explains the next. Under load the rolling elements produce "
            f"a Hertzian contact stress field whose maximum shear stress sits a short distance "
            f"below the raceway surface, not at it. Repeated passes drive microstructural change "
            f"at that depth, and micro cracks initiate there. Those cracks propagate toward the "
            f"surface, and when they break through, a flake of material detaches and leaves a "
            f"pit. That first pit is the important event, because the failure is self "
            f"accelerating from that moment on. The pit destroys the local contact geometry, so "
            f"the rolling element now strikes an edge instead of rolling over a continuous "
            f"curve. The impact raises the local contact stress far above nominal, which "
            f"initiates further cracks at the pit boundary, which enlarges the pit. Detached "
            f"debris also circulates in the oil and work hardens or dents other surfaces, "
            f"seeding new initiation sites elsewhere. In {gbx['tag']} the debris had already "
            f"dented the downwind planet raceway, which is why the whole unit was replaced "
            f"rather than the failed bearing alone when it came out of {subject['tag']}. The "
            f"process is not linear in time: it is slow and nearly invisible until the first "
            f"through crack, then progressively faster. Any monitoring scheme that only watches "
            f"an overall level is watching the wrong part of that curve, because the overall "
            f"level does not move appreciably until the damage is already extensive, and that "
            f"is precisely the trap {std['tag']} is written to avoid.",
        ),
        (
            "Why the alarm fired when it did",
            f"Spalling has a distinctive spectral fingerprint, and understanding it explains "
            f"both the two week warning and the previous months of silence. Each time a rolling "
            f"element passes over the damaged zone it produces a short impulse. Those impulses "
            f"repeat at the bearing pass frequency, which is set by geometry and shaft speed, "
            f"and they amplitude modulate the much stronger gear mesh vibration. In the "
            f"frequency domain, amplitude modulation does not appear as a peak at the "
            f"modulating frequency. It appears as pairs of sidebands straddling the carrier, "
            f"offset by the modulating frequency, which here means pairs around the gear mesh "
            f"frequency spaced at the bearing pass frequency. The figure below shows the "
            f"measured spectrum for {gbx['tag']} on 12 March. The dominant peak is the gear "
            f"mesh frequency at 312 Hz, and it is flanked by symmetric sidebands spaced at "
            f"14.2 Hz, the calculated bearing pass frequency for this unit; the first, second "
            f"and third order sidebands are all present, with amplitude falling as the order "
            f"rises. That pattern is diagnostic. Broadband energy alone was still modest when "
            f"those sidebands were already unambiguous, which is exactly why {std['tag']} "
            f"defines the alarm on the sideband criterion in addition to an overall level, and "
            f"why the overall level alone gave so little notice.",
        ),
        (
            "Contributing factor: the casting batch",
            f"The failure occurred well short of design life, so the analysis looked for a "
            f"contributing factor rather than treating it as a random event. Material "
            f"certificates trace the planet carrier and raceway of {gbx['tag']} to "
            f"{suspect['tag']}, {suspect['foundry']}. Hardness testing of the retained sample "
            f"for that batch returned values at the low end of the specified range, and "
            f"coupons taken from the failed raceway itself confirmed the same. Raceway hardness "
            f"is not a cosmetic property here. It sets the depth and magnitude of the "
            f"subsurface shear stress the material tolerates before micro cracks initiate, so a "
            f"lower hardness shortens the initiation phase described above. It does not change "
            f"the mechanism and it does not change the spectral signature; it changes the clock. "
            f"The conclusion is that {gbx['tag']} failed by an ordinary and well understood "
            f"mechanism on an accelerated schedule, and the accelerant is a material property "
            f"shared by every unit from the same casting run. That is the finding with fleet "
            f"consequences, because {suspect['tag']} was not a single unit purchase. This report "
            f"deliberately does not enumerate the other affected units: the asset register is "
            f"the authoritative record of batch membership and it should be queried rather than "
            f"transcribed into a report that will age.",
        ),
        (
            "Corrective and preventive actions",
            f"Immediate: {gbx['tag']} is removed and replaced, and the replacement unit is "
            f"drawn from a different casting run. The oil system was flushed and the filters "
            f"replaced, because circulating debris from a spalled raceway is an initiation "
            f"source for the replacement unit if it is left in place. Short term: every gearbox "
            f"traceable to {suspect['tag']} is placed on a shortened inspection interval with "
            f"spectral review rather than overall level review, and the sideband criterion in "
            f"{std['tag']} is applied explicitly at each review. Turbine {subject['tag']} was "
            f"not the only unit from that run and the register should be consulted to establish "
            f"the current list, including units whose overall readings are still nominal, since "
            f"this report has established that a nominal overall level is not evidence of a "
            f"healthy raceway. Longer term: incoming hardness verification is added as a "
            f"receiving inspection step for gearbox castings rather than relying on the "
            f"supplier certificate alone, and the condition monitoring configuration is updated "
            f"so that the appearance of pass frequency sidebands raises a notification in its "
            f"own right, independently of the broadband trend.",
        ),
    ]


def build_rca(fleet: dict, out_path: Path, figure_path: str) -> None:
    std = fleet["standard"]
    subject = next(t for t in fleet["turbines"] if t["tag"] == "WTG_A17")
    gbx = subject["gearbox"]
    signer = next(t for t in fleet["technicians"] if t["signsDocuments"])

    styles = getSampleStyleSheet()
    body = ParagraphStyle(
        "NwBody",
        parent=styles["BodyText"],
        fontSize=9.5,
        leading=14,
        alignment=TA_JUSTIFY,
        spaceAfter=8,
    )
    heading = ParagraphStyle(
        "NwHeading", parent=styles["Heading2"], fontSize=12, spaceBefore=12, spaceAfter=6
    )

    story: list = [
        Paragraph(
            f"Root cause analysis RCA_2026_014: high speed stage failure of {gbx['tag']}",
            ParagraphStyle("NwTitle", parent=styles["Title"], fontSize=16, leading=20),
        ),
        Paragraph(
            # Dated AFTER the events it reports: the alarm fired 14 March and this report already
            # contains strip-down findings and hardness coupon results, which take weeks.
            f"{fleet['operator']} | asset {subject['tag']} | prepared by {signer['name']} | "
            f"10 April 2026 | classification: internal | synthetic sample document",
            ParagraphStyle("NwMeta", parent=styles["Normal"], fontSize=8,
                           textColor=colors.HexColor("#5A6068"), spaceAfter=10),
        ),
    ]

    for index, (title, text) in enumerate(rca_sections(fleet)):
        story.append(Paragraph(title, heading))
        story.append(Paragraph(text, body))
        # The figure belongs to the "Why the alarm fired" section, which is index 3.
        if index == 3:
            story.append(Spacer(1, 0.2 * cm))
            story.append(Image(figure_path, width=16 * cm, height=7.03 * cm))
            story.append(
                Paragraph(
                    f"Figure 1. Velocity spectrum at the high speed stage of {gbx['tag']}, "
                    f"12 March 2026. The dashed line is the {std['tag']} BROADBAND alarm limit of "
                    f"{std['alarmMmPerS']} mm/s RMS, shown for reference only: it applies to the "
                    f"quadrature sum across the band, not to any single spectral line. Note that "
                    f"no line approaches it while the sideband family is already unambiguous.",
                    ParagraphStyle("NwCaption", parent=styles["Italic"], fontSize=8,
                                   textColor=colors.HexColor("#5A6068"), spaceBefore=4,
                                   spaceAfter=10),
                )
            )

    SimpleDocTemplate(
        str(out_path),
        pagesize=A4,
        title=f"RCA_2026_014 {gbx['tag']}",
        author=fleet["operator"],
        subject="Gearbox high speed stage failure",
        leftMargin=2 * cm,
        rightMargin=2 * cm,
        topMargin=1.8 * cm,
        bottomMargin=1.8 * cm,
        invariant=1,
    ).build(story)


# --------------------------------------------------------------------------------------
# Document 2: the maintenance register (XLSX, the table document)
# --------------------------------------------------------------------------------------
def build_register(fleet: dict, out_path: Path) -> None:
    std = fleet["standard"]
    signer = next(t for t in fleet["technicians"] if t["signsDocuments"])
    by_tag = {t["tag"]: t for t in fleet["turbines"]}

    wb = Workbook()
    ws = wb.active
    ws.title = "Register"

    header = [
        "Asset tag", "Gearbox", "Casting batch", "Substation", "Last service",
        "Vibration mm/s RMS", "Status", "Technician", "Work order", "Vendor",
    ]
    ws.append(header)
    for tag in fleet["registerRows"]:
        turbine = by_tag[tag]
        gbx = turbine["gearbox"]
        vendor = fleet["vendors"][0] if turbine["site"] == "SITE_NORD_ESBJERG" else fleet["vendors"][1]
        ws.append([
            turbine["tag"], gbx["tag"], gbx["batch"], turbine["substation"],
            gbx["lastService"], gbx["vibration"], gbx["status"], gbx["technician"],
            gbx["workOrder"], vendor,
        ])

    bold = Font(bold=True)
    fill = PatternFill("solid", fgColor="E8EAEC")
    for cell in ws[1]:
        cell.font = bold
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    widths = [12, 13, 20, 14, 13, 12, 11, 18, 12, 16]
    for column, width in zip("ABCDEFGHIJ", widths):
        ws.column_dimensions[column].width = width
    ws.freeze_panes = "A2"

    # A prose sheet so the spreadsheet also contributes named entities, and so the reader
    # knows what the register is and is NOT. Kept to single cells per line.
    notes = wb.create_sheet("Notes")
    for line in [
        [f"{fleet['operator']} gearbox maintenance register, extract of recently serviced units."],
        [f"Site: Esbjerg Nord and Horns Vest, {fleet['region']}. Grid connection {fleet['grid']['tag']}."],
        [f"Vibration readings are broadband velocity at the high speed stage, measured under "
         f"the method in standard {std['tag']} ({std['revision']})."],
        [f"The alarm limit is {std['alarmMmPerS']} mm/s RMS and the warning level is "
         f"{std['warningMmPerS']} mm/s RMS."],
        [f"Gearboxes are supplied by {fleet['vendors'][0]} and {fleet['vendors'][1]}. "
         f"Strip down inspections are witnessed by the supplier."],
        [f"This extract lists recently serviced units only. It is NOT a complete asset "
         f"register and it is NOT a complete record of casting batch membership; query the "
         f"asset register for those."],
        [f"Extract approved by {signer['name']}, {signer['role'].lower()}, on 20 March 2026."],
    ]:
        notes.append(line)
    notes.column_dimensions["A"].width = 118
    for row in notes.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.properties.creator = fleet["operator"]
    wb.properties.title = "Gearbox maintenance register"
    wb.properties.created = FIXED_TIME
    wb.properties.modified = FIXED_TIME
    wb.save(out_path)
    normalize_xlsx(out_path)


# --------------------------------------------------------------------------------------
# Document 3: the engineering standard (markdown, text only, no sidecar)
# --------------------------------------------------------------------------------------
def standard_sections(fleet: dict) -> list[tuple[str, str]]:
    std = fleet["standard"]
    signer = next(t for t in fleet["technicians"] if t["signsDocuments"])
    analyst = next(t for t in fleet["technicians"] if t["role"].startswith("Condition"))
    alarm = std["alarmMmPerS"]
    warning = std["warningMmPerS"]

    return [
        (
            "Scope and normative references",
            f"This standard defines how {fleet['operator']} monitors gearbox vibration on its "
            f"offshore fleet in {fleet['region']}, and it sets the levels at which action is "
            f"required. It applies to every geared turbine at the Esbjerg Nord and Horns Vest "
            f"sites, whether supplied by {fleet['vendors'][0]} or {fleet['vendors'][1]}, and it "
            f"applies to both permanently installed condition monitoring and to handheld route "
            f"measurements. It does not cover direct drive machines, which have no gearbox and "
            f"are governed separately, and it does not cover generator or blade bearing "
            f"monitoring. The velocity levels here are expressed as broadband RMS in millimetres "
            f"per second, measured over the band defined in the measurement section below, and "
            f"they are consistent with the general severity guidance for this machine class in "
            f"ISO 10816. Where this standard is stricter than the turbine supplier's own "
            f"published limits, this standard governs, because the supplier limits are written "
            f"for warranty purposes and this standard is written to catch damage early enough "
            f"to plan a vessel. This revision is {std['revision']}, approved by "
            f"{signer['name']}, {signer['role'].lower()}.",
        ),
        (
            "Why a single overall level is not sufficient",
            f"An overall broadband level is a single number summarising all the vibration energy "
            f"in the measured band, and it is genuinely useful because it is cheap, stable and "
            f"trendable. It is also, on its own, a poor detector of the most common serious "
            f"gearbox fault, and the reason is arithmetic rather than opinion. Bearing raceway "
            f"damage in its early stage produces short, low energy impulses. Spread across a "
            f"wide measurement band and combined in quadrature with the much larger gear mesh "
            f"and shaft components that are always present, those impulses contribute very "
            f"little to the total. A raceway can carry a through crack and a growing pit while "
            f"the overall level moves by a few tenths of a millimetre per second, which is "
            f"within the range that normal load and temperature variation produces anyway. By "
            f"the time the damage is extensive enough to lift the overall level clearly, the "
            f"failure is in its self accelerating phase and the remaining useful life is measured "
            f"in days. Setting the overall alarm lower does not solve this: it produces false "
            f"alarms from benign operating variation without materially improving early "
            f"detection. The answer is not a lower threshold on the same number. It is a second "
            f"criterion that looks at the shape of the spectrum instead of its total.",
        ),
        (
            "The sideband criterion",
            f"Early raceway damage announces itself by modulation rather than by magnitude. Each "
            f"pass of a rolling element over a damaged zone produces an impulse, those impulses "
            f"repeat at the bearing pass frequency determined by geometry and shaft speed, and "
            f"they amplitude modulate the strong gear mesh vibration that is always present. "
            f"Amplitude modulation in the frequency domain does not appear as a peak at the "
            f"modulating frequency, which is the point most often missed: it appears as pairs of "
            f"sidebands straddling the carrier, offset above and below it by the modulating "
            f"frequency. So the diagnostic pattern is symmetric sidebands around the gear mesh "
            f"frequency, spaced at the calculated bearing pass frequency, typically with first, "
            f"second and third order pairs visible and amplitude falling as order rises. This "
            f"pattern is specific: normal gear wear raises the mesh peak and its harmonics "
            f"without producing pass frequency sidebands, and looseness or misalignment produces "
            f"shaft order families instead. Because the pattern is specific and appears early, "
            f"it is the primary criterion in this standard. The presence of a clear pass "
            f"frequency sideband family requires spectral review and a shortened inspection "
            f"interval regardless of the overall level, and it does so on its own authority: a "
            f"nominal overall reading is not grounds to dismiss it.",
        ),
        (
            "Levels and why they sit where they do",
            f"Three levels apply to the broadband velocity at the high speed stage. Below "
            f"{warning} mm/s RMS the unit is nominal and the normal route interval applies. At "
            f"or above {warning} mm/s RMS the unit is at warning: the spectrum is reviewed by "
            f"the condition monitoring analyst, currently {analyst['name']}, the trend rate is "
            f"assessed, and the inspection interval is shortened. At or above {alarm} mm/s RMS "
            f"the unit is in alarm: the turbine is stopped and the drivetrain is inspected "
            f"before it returns to service. The alarm level is not derived from a severity chart "
            f"alone. It is set where the expected remaining life at the observed progression "
            f"rate for this machine class still leaves time to mobilise a vessel and a crane, "
            f"because a limit that is exceeded only when repair is already impossible to plan "
            f"has no operational value. That reasoning is also why the alarm sits above the "
            f"warning level by a deliberately narrow margin: the interval between them is "
            f"intended to be short in time, and the warning is where the useful work happens. "
            f"Treating the alarm as the real threshold and the warning as advisory inverts the "
            f"intent of this standard.",
        ),
        (
            "Measurement method and record keeping",
            f"Measurements are taken at the high speed stage bearing housing in the radial "
            f"direction, at the permanently marked location, with the machine at or above "
            f"seventy percent of rated power and with the drivetrain at operating temperature. "
            f"Readings taken below that load or on a cold drivetrain are not comparable to the "
            f"trend and are recorded as indicative only. The broadband figure is the RMS "
            f"velocity from 10 Hz to 1000 Hz. Spectral review additionally requires a resolution "
            f"fine enough to separate sidebands spaced at the calculated bearing pass frequency "
            f"from the gear mesh peak, which in practice means a line spacing several times "
            f"finer than that spacing; a spectrum too coarse to resolve the sidebands cannot be "
            f"used to dismiss the sideband criterion, and reviewing one is a recorded "
            f"non conformance. Every measurement is recorded against the asset tag of the "
            f"gearbox, not the turbine, so that a unit which moves between positions keeps its "
            f"own history and a replacement unit does not inherit the history of the one it "
            f"replaced. Where a reading leads to an intervention, the work order reference is "
            f"recorded alongside it, and where a strip down establishes a mechanism, the root "
            f"cause report reference is recorded as well.",
        ),
    ]


def build_standard(fleet: dict, out_path: Path) -> None:
    std = fleet["standard"]
    signer = next(t for t in fleet["technicians"] if t["signsDocuments"])

    lines = [
        f"# {std['tag']}: {std['title']}",
        "",
        f"{fleet['operator']} engineering standard | {std['revision']} | "
        f"approved by {signer['name']} | effective 1 April 2026",
        "",
    ]
    for title, text in standard_sections(fleet):
        lines.append(f"## {title}")
        lines.append("")
        lines.append(text)
        lines.append("")

    out_path.write_text("\n".join(lines), encoding="utf-8", newline="\n")


def main() -> None:
    fleet = load_fleet()
    with tempfile.TemporaryDirectory() as scratch:
        figure = os.path.join(scratch, "vibration-spectrum.png")
        build_figure(figure, fleet["standard"]["alarmMmPerS"])
        build_rca(fleet, HERE / "nw-rca-wtg-a17.pdf", figure)
    build_register(fleet, HERE / "nw-fleet-register.xlsx")
    build_standard(fleet, HERE / "nw-std-0417.md")

    for name in ("nw-rca-wtg-a17.pdf", "nw-fleet-register.xlsx", "nw-std-0417.md"):
        size = (HERE / name).stat().st_size
        print(f"  wrote samples/documents/{name} ({size / 1024:.1f} KiB)")

    # Section lengths decide chunk boundaries: anything at or below 800 characters merges
    # into its neighbour and the chunk chain collapses. Report them so a content edit that
    # breaks the demo is visible here rather than after an ingest.
    print("\n  section lengths (must exceed 800 characters to stay separate chunks):")
    for label, sections in (("rca", rca_sections(fleet)),
                            ("std", standard_sections(fleet))):
        for title, text in sections:
            flag = "ok " if len(text) > 800 else "SHORT"
            print(f"    [{flag}] {label}/{title}: {len(text)}")


if __name__ == "__main__":
    main()
